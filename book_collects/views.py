# -*- coding: utf-8 -*-

from django.http import HttpResponse, HttpResponseRedirect
from django.shortcuts import render_to_response
from django.views.decorators.csrf import csrf_exempt
from django.contrib import auth
from models import *
from openpyxl import load_workbook
import os
import uuid
import cookielib
import urllib2
import time
import re
import traceback
import BeautifulSoup
import HTMLParser
import threading
import Queue

global result_num

html_parser = HTMLParser.HTMLParser()

cj = cookielib.CookieJar()
#cj = cookielib.LWPCookieJar()
opener = urllib2.build_opener(urllib2.HTTPCookieProcessor(cj))
opener.addheaders = [('User-agent', 'Mozilla/5.0 (Windows NT 5.2) AppleWebKit/537.1 (KHTML, like Gecko) Chrome/21.0.1180.89 Safari/537.1'),
                     ('Accept', 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8'),
                     ('Accept-Language', 'zh-cn,zh;q=0.8,en-us;q=0.5,en;q=0.3'),
                     ('Connection', 'keep-alive')
]
opener.addheaders.append( ('Accept-encoding', 'identity') )
opener.addheaders.append( ('Referer', '') )

def get_page(url, data=None):
    print "Opening:", url
    n = 0
    while n < 5:
        n = n + 1
        try:
            resp = opener.open(url, data, timeout=20)
            page = resp.read()
            return page
        except:
            print "retry", n
            traceback.print_exc()
            print "Will try after 2 seconds ..."
            time.sleep(2.0)
            continue
        break
    return ""


def add_art(title, book_name, page, author_name, mail, link):
    title = html_parser.unescape(title)
    book_name = html_parser.unescape(book_name)
    page = html_parser.unescape(page)
    author_name = html_parser.unescape(author_name)
    mail = html_parser.unescape(mail)

    chars = "ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz"
    while True:
        print author_name
        if not author_name:
            break
        if author_name[-1] in chars:
            break
        else:
            author_name = author_name[:-1]

    art = Art()
    art.title = title
    art.book_name = book_name
    art.page = page
    art.author_name = author_name
    art.mail = mail
    art.link = link
    art.save()

    print title
    print book_name
    print page
    print author_name
    print mail
    print link
    print "-------------------"


def add_read(issue_links):
    for issue_url in issue_links:
        if not Read.objects.filter(url=issue_url):
            r = Read()
            r.url = issue_url
            r.save()



class Collect_issue(threading.Thread):
    def __init__(self,issue_queue,queue):
        self.queue=queue
        self.issue_queue=issue_queue
        threading.Thread.__init__(self)
    def run(self):
        while not self.issue_queue.empty():
            link = self.issue_queue.get()
            if "sciencedirect" in link:
                i_str = get_page(link)
                ts = re.findall(r'<a.*?href="(.*?)".*?artTitle.*?</a>', i_str)
                for t in ts:
                    if t not in self.queue.queue:
                        self.queue.put(t)
            elif "onlinelibrary" in link:
                if Read.objects.filter(url=link):
                    continue
                i_str = get_page(link)
                ts = re.findall(r'<div class="citation tocArticle"><a href="(.*?)"', i_str)
                for t in ts:
                    if t[0] == "/":
                        t = "http://onlinelibrary.wiley.com" + t
                    if t not in self.queue.queue:
                        self.queue.put(t)




class Collect_link(threading.Thread):
    def __init__(self,queue):
        self.queue=queue
        threading.Thread.__init__(self)
    def run(self):
        global result_num
        while not self.queue.empty():
            link = self.queue.get()

            if Art.objects.filter(link=link):
                continue

            try:
                if "sciencedirect" in link:
                    url = link + "?np=y"
                    p_str = get_page(url)

                    p_soup = BeautifulSoup.BeautifulSoup(p_str)
                    title = p_soup.find("title").getText()
                    book_name = p_soup.find("div", "title").getText()
                    page = p_soup.find("p", "volIssue").getText().replace(u"â&euro;&ldquo;", " - ")

                    authorgroup = p_soup.find("ul", "authorGroup")
                    if authorgroup:
                        lis = authorgroup.findAll("li")
                        for li in lis:
                            author_name = li.find("a", "authorName").getText()
                            mails = re.findall(r'href="mailto:(.*?)"', str(li))
                            for mail in mails:
                                result_num += 1
                                add_art(title, book_name, page, author_name, mail, link)

                elif "onlinelibrary" in link:
                    p_str = get_page(link)
                    p_soup = BeautifulSoup.BeautifulSoup(p_str)
                    title = p_soup.find("span", "mainTitle").getText()
                    book_name = p_soup.find(id="productTitle").getText()
                    page = p_soup.find("p", "articleDetails").getText()

                    if p_soup.find(id="authors"):
                        author = ""
                        author_lis = p_soup.find(id="authors").findAll("li")
                        for li in author_lis:
                            if "*" in str(li):
                                li_text = li.getText()
                                ti = li_text.find("*")
                                author = li_text[:ti]
                                while True:
                                    if not author:
                                        break
                                    if author[0] in "ABCDEFGHIJKLMNOPQRSTUVWXYZ":
                                        break
                                    else:
                                        author = author[1:]
                        mails = re.findall(r'href="mailto:(.*?)"', p_str)
                        for mail in mails:
                            mail = mail.replace("%E2%80%90", "-")
                            author_name = author
                            if not author:
                                author_name = mail[:mail.find("@")].replace(".", " ")
                            result_num += 1
                            add_art(title, book_name, page, author_name, mail, link)
                        if not mails:
                            span_email = p_soup.find("span", "email")
                            if span_email:
                                t = span_email.getText()
                                author_name = t[:t.find("(")]
                                mail = t[t.find("(")+1:-1]
                                add_art(title, book_name, page, author_name, mail, link)

            except:
                print link, "+"





def collect_sciencedirect(url):
    print "===========", url, "============="

    try:
        host_url = 'http://www.sciencedirect.com'
        #url = 'http://www.sciencedirect.com/science/journal/19389736'
        p_str = get_page(url)
        p_soup = BeautifulSoup.BeautifulSoup(p_str)

        year_links = []
        issue_links = []
        links = []

        tab = p_soup.find(id="volumeIssueData")

        year_as = tab.findAll("a", {"aria-expanded":"true"})
        year_as += tab.findAll("a", {"aria-expanded":"false"})

        for year_a in year_as:
            t = year_a.get("href")
            if t[0] == "/":
                t = host_url + t
            if t not in year_links:
                year_links.append(t)




        for year_link in year_links:
            if Read.objects.filter(url=year_link):
                continue
            issue_links.append(year_link)
            y_str = get_page(year_link)
            y_soup = BeautifulSoup.BeautifulSoup(y_str)
            tab = y_soup.find(id="volumeIssueData")
            issue_divs = tab.findAll("div", "txt")
            for div in issue_divs:
                issue_link = div.find("a")
                if issue_link:
                    issue_link = issue_link.get("href")
                if issue_link and issue_link[0] == "/":
                    issue_link = host_url + issue_link
                if issue_link and issue_link not in issue_links:
                    issue_links.append(issue_link)


        issue_queue = Queue.Queue()
        queue = Queue.Queue()

        for link in issue_links:
            issue_queue.put(link)

        t_num = 10
        threads=[]
        for i in xrange(t_num):
            threads.append(Collect_issue(issue_queue, queue))
        for i in xrange(t_num):
            threads[i].start()
        for i in xrange(t_num):
            threads[i].join()


        print queue.queue


        t_num = 10
        threads=[]
        for i in xrange(t_num):
            threads.append(Collect_link(queue))
        for i in xrange(t_num):
            threads[i].start()
        for i in xrange(t_num):
            threads[i].join()

        print "finish"
        add_read(year_links[1:])

    except:
        pass



def collect_onlinelibrary(url):
    print "===========", url, "============="

    try:

        host_url = "http://onlinelibrary.wiley.com"
        #url = 'http://onlinelibrary.wiley.com/journal/10.1111/(ISSN)1463-6395'
        links = []
        issue_links =[]
        for year in range(2000,2015):
            print year
            year_url = url + "/issues/fragment?activeYear=" + str(year)
            p_str = get_page(year_url)
            issues = re.findall(r'<div class="issue"><a href="(.*?)"', p_str)

            for issue in issues:
                if issue[0] == "/":
                    issue = host_url + issue
                if issue not in issue_links:
                    issue_links.append(issue)

        print issue_links

        issue_queue = Queue.Queue()
        queue = Queue.Queue()

        for link in issue_links:
            issue_queue.put(link)

        t_num = 10
        threads=[]
        for i in xrange(t_num):
            threads.append(Collect_issue(issue_queue, queue))
        for i in xrange(t_num):
            threads[i].start()
        for i in xrange(t_num):
            threads[i].join()


        print queue.queue


        t_num = 10
        threads=[]
        for i in xrange(t_num):
            threads.append(Collect_link(queue))
        for i in xrange(t_num):
            threads[i].start()
        for i in xrange(t_num):
            threads[i].join()

        print "finish"
        add_read(issue_links)

    except:
        pass







def index(request):
    wb = load_workbook(filename='books.xlsx')
    sheet_name = wb.get_sheet_names()[0]
    sheet = wb.get_sheet_by_name(sheet_name)
    urls = ""
    n = 0
    while True:
        n += 1
        if not sheet.cell("B%d"%n).value:
            break
        # name = str(sheet.cell("A%d"%n).value.encode("gbk"))
        url = str(sheet.cell("B%d"%n).value)
        urls += "%s\n" % url
    return render_to_response('index.html', locals())


def collect(request):
    global result_num
    result_num = 0
    urls = request.REQUEST.get("urls")
    urls = urls.split("\n")
    for url in urls:
        url = url.split(",")[-1].strip()
        if not url:
            continue
        if Collected.objects.filter(url=url).count():
            continue
        if "sciencedirect" in url:
            collect_sciencedirect(url)
        elif "onlinelibrary" in url:
            collect_onlinelibrary(url)
        Collected(url=url).save()
    time.sleep(5)
    Collected.objects.all().delete()
    return HttpResponse(str(result_num))


def data(request):
    arts = Art.objects.all()
    return render_to_response('data.html', locals())


def del_art(request, id):
    Art.objects.filter(id=id).delete()
    return HttpResponseRedirect("/data/")


def update_art(request):
    id = request.REQUEST.get("id")
    title = request.REQUEST.get("title")
    book_name = request.REQUEST.get("book_name")
    page = request.REQUEST.get("page")
    author_name = request.REQUEST.get("author_name")
    mail = request.REQUEST.get("mail")
    link = request.REQUEST.get("link")
    art = Art.objects.get(id=id)
    art.title = title
    art.book_name = book_name
    art.author_name = author_name
    art.page = page
    art.mail = mail
    art.link = link
    art.save()
    return HttpResponseRedirect("/data/")

